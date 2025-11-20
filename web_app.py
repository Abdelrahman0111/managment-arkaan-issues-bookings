from flask import Flask, render_template, request, jsonify, send_file
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from datetime import datetime
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

app = Flask(__name__)

# الاتصال بـ Google Sheets
_cached_spreadsheet = None
_connection_error = None

def connect_to_gsheet():
    global _cached_spreadsheet, _connection_error
    
    if _connection_error:
        return None
    
    if _cached_spreadsheet:
        return _cached_spreadsheet
    
    try:
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file('credentials.json', scopes=scope)
        client = gspread.authorize(creds)
        _cached_spreadsheet = client.open('نظام_المشاكل')
        return _cached_spreadsheet
    except Exception as e:
        _connection_error = str(e)
        print(f"Error connecting to Google Sheets: {e}")
        print("\n⚠️ يرجى إنشاء Service Account جديد من Google Cloud Console")
        return None

# جلب بيانات المشاكل
def get_issues():
    spreadsheet = connect_to_gsheet()
    if spreadsheet:
        try:
            sheet = spreadsheet.worksheet('المشاكل')
            return sheet.get_all_records()
        except:
            sheet = spreadsheet.add_worksheet(title='المشاكل', rows=1000, cols=15)
            sheet.append_row(['agent_name', 'booking_number', 'discount', 'notes', 'check_in', 'check_out', 'created_at', 'issue_type', 'payment_type', 'monthly_amount', 'paid_amount', 'remaining_amount', 'payment_status'])
            return []
    return []

# جلب بيانات الوكلاء
def get_agents():
    spreadsheet = connect_to_gsheet()
    if spreadsheet:
        try:
            sheet = spreadsheet.worksheet('الوكلاء')
            return sheet.get_all_records()
        except:
            sheet = spreadsheet.add_worksheet(title='الوكلاء', rows=1000, cols=5)
            sheet.append_row(['agent_name', 'created_at'])
            return []
    return []

@app.route('/')
def dashboard():
    issues = get_issues()
    agents = get_agents()
    
    stats = {
        'total_issues': len(issues),
        'total_agents': len(set([i['agent_name'] for i in issues])) if issues else 0,
        'simple_issues': len([i for i in issues if i.get('issue_type') == 'مشكلة بسيطة']),
        'major_issues': len([i for i in issues if i.get('issue_type') == 'مشكلة كبيرة'])
    }
    
    return render_template('dashboard.html', stats=stats, recent_issues=issues[-5:][::-1] if issues else [])

@app.route('/agents')
def agents_page():
    agents = get_agents()
    return render_template('agents.html', agents=agents)

@app.route('/add_agent', methods=['POST'])
def add_agent():
    agent_name = request.json.get('agent_name')
    if agent_name:
        spreadsheet = connect_to_gsheet()
        if spreadsheet:
            sheet = spreadsheet.worksheet('الوكلاء')
            created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sheet.append_row([agent_name, created_at])
            return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/issues')
def issues_page():
    agents = get_agents()
    return render_template('add_issue.html', agents=[a['agent_name'] for a in agents])

@app.route('/add_issue', methods=['POST'])
def add_issue():
    data = request.json
    spreadsheet = connect_to_gsheet()
    if spreadsheet:
        sheet = spreadsheet.worksheet('المشاكل')
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        discount = float(data.get('discount', 0))
        payment_type = data.get('payment_type', 'كلي')
        monthly_amount = float(data.get('monthly_amount', 0)) if payment_type == 'جزئي' else discount
        paid_amount = 0
        remaining_amount = discount
        payment_status = 'لم يدفع' if discount > 0 else 'مكتمل'
        
        row = [
            data['agent_name'], data['booking_number'], discount,
            data['notes'], data['check_in'], data['check_out'], created_at, data['issue_type'],
            payment_type, monthly_amount, paid_amount, remaining_amount, payment_status
        ]
        sheet.append_row(row)
        return jsonify({'success': True})
    return jsonify({'success': False})

@app.route('/view_issues')
def view_issues():
    issues = get_issues()
    print(f"Issues found: {len(issues)}")  # للتأكد
    return render_template('view_issues.html', issues=issues)

@app.route('/payments')
def payments():
    issues = get_issues()
    return render_template('payments.html', issues=issues)

@app.route('/record_payment', methods=['POST'])
def record_payment():
    data = request.json
    booking_number = data.get('booking_number')
    payment_amount = float(data.get('payment_amount', 0))
    
    spreadsheet = connect_to_gsheet()
    if spreadsheet:
        sheet = spreadsheet.worksheet('المشاكل')
        all_data = sheet.get_all_values()
        
        for idx, row in enumerate(all_data[1:], start=2):
            if row[1] == booking_number:
                discount = float(row[2])
                paid_amount = float(row[10]) if row[10] else 0
                new_paid = paid_amount + payment_amount
                remaining = discount - new_paid
                
                status = 'مكتمل' if remaining <= 0 else 'جزئي'
                
                sheet.update_cell(idx, 11, new_paid)
                sheet.update_cell(idx, 12, max(0, remaining))
                sheet.update_cell(idx, 13, status)
                
                return jsonify({'success': True})
    
    return jsonify({'success': False}), 400

@app.route('/reports')
def reports():
    issues = get_issues()
    agents = list(set([i['agent_name'] for i in issues])) if issues else []
    return render_template('reports.html', agents=agents)

@app.route('/export_excel', methods=['POST'])
def export_excel():
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    issues = get_issues()
    if not issues:
        return jsonify({'error': 'لا توجد بيانات'}), 400
    
    df = pd.DataFrame(issues)
    
    # تحويل التواريخ للعربية
    arabic_months = {
        '01': 'يناير', '02': 'فبراير', '03': 'مارس', '04': 'أبريل',
        '05': 'مايو', '06': 'يونيو', '07': 'يوليو', '08': 'أغسطس',
        '09': 'سبتمبر', '10': 'أكتوبر', '11': 'نوفمبر', '12': 'ديسمبر'
    }
    
    def format_arabic_date(date_str):
        try:
            if pd.isna(date_str) or not date_str:
                return ''
            date_obj = pd.to_datetime(date_str)
            day = date_obj.day
            month = arabic_months[date_obj.strftime('%m')]
            year = date_obj.year
            return f"{day} {month} {year}"
        except:
            return str(date_str)
    
    df['check_in'] = df['check_in'].apply(format_arabic_date)
    df['check_out'] = df['check_out'].apply(format_arabic_date)
    df['created_at'] = df['created_at'].apply(format_arabic_date)
    df['discount'] = pd.to_numeric(df['discount'], errors='coerce').fillna(0)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # صفحة المشاكل التفصيلية
        df_display = df[['agent_name', 'booking_number', 'issue_type', 'discount', 'notes', 'check_in', 'check_out', 'created_at']]
        df_display.columns = ['اسم الوكيل', 'رقم الحجز', 'نوع المشكلة', 'الخصم', 'الملاحظات', 'تسجيل الدخول', 'تسجيل الخروج', 'تاريخ الإنشاء']
        
        df_display.to_excel(writer, sheet_name='تفاصيل المشاكل', index=False, startrow=3)
        ws = writer.sheets['تفاصيل المشاكل']
        
        # عنوان التقرير
        ws['A1'] = 'تقرير المشاكل التفصيلي'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f'تاريخ التقرير: {format_arabic_date(datetime.now())}'
        ws['A2'].font = Font(size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:H2')
        
        # تنسيق الهيدر
        header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 9):
            cell = ws.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # تنسيق البيانات
        for row in range(5, len(df_display) + 5):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 4:  # عمود الخصم
                    cell.number_format = '#,##0.00'
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        
        # صلحة الملخص العام
        summary_data = [
            ['البيان', 'العدد'],
            ['إجمالي المشاكل', len(df)],
            ['مشاكل بسيطة', len(df[df['issue_type'] == 'مشكلة بسيطة'])],
            ['مشاكل متوسطة', len(df[df['issue_type'] == 'مشكلة متوسطة'])],
            ['مشاكل كبيرة', len(df[df['issue_type'] == 'مشكلة كبيرة'])],
            ['إجمالي الخصومات', df['discount'].sum()]
        ]
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='الملخص العام', index=False, header=False, startrow=2)
        ws_summary = writer.sheets['الملخص العام']
        
        ws_summary['A1'] = 'الملخص العام'
        ws_summary['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws_summary['A1'].fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        ws_summary['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_summary.merge_cells('A1:B1')
        
        for row in range(3, 9):
            for col in range(1, 3):
                cell = ws_summary.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                if col == 2 and row == 8:
                    cell.number_format = '#,##0.00'
        
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        
        # صفحة ملخص الوكلاء
        agent_summary = df.groupby('agent_name').agg({
            'booking_number': 'count',
            'discount': 'sum'
        }).reset_index()
        agent_summary.columns = ['اسم الوكيل', 'عدد المشاكل', 'إجمالي الخصومات']
        
        agent_summary.to_excel(writer, sheet_name='ملخص الوكلاء', index=False, startrow=2)
        ws_agents = writer.sheets['ملخص الوكلاء']
        
        ws_agents['A1'] = 'ملخص الخصومات حسب الوكيل'
        ws_agents['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws_agents['A1'].fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        ws_agents['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_agents.merge_cells('A1:C1')
        
        for col in range(1, 4):
            cell = ws_agents.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        for row in range(4, len(agent_summary) + 4):
            for col in range(1, 4):
                cell = ws_agents.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 3:
                    cell.number_format = '#,##0.00'
        
        ws_agents.column_dimensions['A'].width = 25
        ws_agents.column_dimensions['B'].width = 15
        ws_agents.column_dimensions['C'].width = 20
    
    output.seek(0)
    return send_file(output, download_name=f'تقرير_مفصل_{datetime.now().strftime("%Y%m%d")}.xlsx', as_attachment=True)

@app.route('/export_pdf', methods=['POST'])
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from bidi.algorithm import get_display
    import arabic_reshaper
    
    issues = get_issues()
    if not issues:
        return jsonify({'error': 'لا توجد بيانات'}), 400
    
    df = pd.DataFrame(issues)
    df['discount'] = pd.to_numeric(df['discount'], errors='coerce').fillna(0)
    
    # تحويل التواريخ للعربية
    arabic_months = {
        '01': 'يناير', '02': 'فبراير', '03': 'مارس', '04': 'أبريل',
        '05': 'مايو', '06': 'يونيو', '07': 'يوليو', '08': 'أغسطس',
        '09': 'سبتمبر', '10': 'أكتوبر', '11': 'نوفمبر', '12': 'ديسمبر'
    }
    
    def format_arabic_date(date_str):
        try:
            if pd.isna(date_str) or not date_str:
                return ''
            date_obj = pd.to_datetime(date_str)
            return f"{date_obj.day} {arabic_months[date_obj.strftime('%m')]} {date_obj.year}"
        except:
            return str(date_str)
    
    def arabic_text(text):
        try:
            reshaped = arabic_reshaper.reshape(str(text))
            return get_display(reshaped)
        except:
            return str(text)
    
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    
    # العنوان
    title_style = ParagraphStyle('Title', fontSize=20, alignment=1, spaceAfter=20)
    elements.append(Paragraph(arabic_text('تقرير المشاكل التفصيلي'), title_style))
    elements.append(Paragraph(arabic_text(f'تاريخ التقرير: {format_arabic_date(datetime.now())}'), title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # جدول الملخص العام
    summary_data = [
        [arabic_text('البيان'), arabic_text('العدد')],
        [arabic_text('إجمالي المشاكل'), str(len(df))],
        [arabic_text('مشاكل بسيطة'), str(len(df[df['issue_type'] == 'مشكلة بسيطة']))],
        [arabic_text('مشاكل متوسطة'), str(len(df[df['issue_type'] == 'مشكلة متوسطة']))],
        [arabic_text('مشاكل كبيرة'), str(len(df[df['issue_type'] == 'مشكلة كبيرة']))],
        [arabic_text('إجمالي الخصومات'), f"{df['discount'].sum():.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # جدول ملخص الوكلاء
    agent_summary = df.groupby('agent_name').agg({
        'booking_number': 'count',
        'discount': 'sum'
    }).reset_index()
    
    agent_data = [[arabic_text('اسم الوكيل'), arabic_text('عدد المشاكل'), arabic_text('إجمالي الخصومات')]]
    for _, row in agent_summary.iterrows():
        agent_data.append([
            arabic_text(row['agent_name']),
            str(row['booking_number']),
            f"{row['discount']:.2f}"
        ])
    
    agent_table = Table(agent_data, colWidths=[3*inch, 2*inch, 2*inch])
    agent_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(agent_table)
    elements.append(PageBreak())
    
    # جدول المشاكل التفصيلية
    elements.append(Paragraph(arabic_text('تفاصيل المشاكل'), title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    issues_data = [[
        arabic_text('الوكيل'),
        arabic_text('رقم الحجز'),
        arabic_text('نوع المشكلة'),
        arabic_text('الخصم'),
        arabic_text('تاريخ الإنشاء')
    ]]
    
    for _, row in df.iterrows():
        issues_data.append([
            arabic_text(row['agent_name'])[:15],
            str(row['booking_number'])[:12],
            arabic_text(row['issue_type'])[:12],
            f"{row['discount']:.2f}",
            format_arabic_date(row['created_at'])[:15]
        ])
    
    issues_table = Table(issues_data, colWidths=[1.5*inch, 1.3*inch, 1.5*inch, 1*inch, 1.5*inch])
    issues_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(issues_table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    
    return send_file(pdf_buffer, download_name=f'تقرير_مفصل_{datetime.now().strftime("%Y%m%d")}.pdf', as_attachment=True, mimetype='application/pdf')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
